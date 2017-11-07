'''

Author: Mason Edgar
Symptom Checking Algorithm for the PMDK
University of Houston | ECE
Senior Capstone - 2017

'''


import numpy as np
import pandas as pd
from openpyxl import load_workbook
import itertools as iter
import Algorithm_1_Method as algo
import xlsxwriter as xlsx


df = pd.read_excel('/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Illness Matrix Barebones .xlsx')
illness_matrix = df.as_matrix()

np.set_printoptions(threshold=np.nan)

Anemia = illness_matrix[0]
Bronchitis = illness_matrix[1]
ColdSore = illness_matrix[2]
Conjunctivitis = illness_matrix[3]
Diabetes = illness_matrix[4]
Chickenpox = illness_matrix[5]
KidneyStones = illness_matrix[6]
Migraines = illness_matrix[7]
PollenAllergy = illness_matrix[8]
Tuberculosis = illness_matrix[9]

AnemiaSymptoms = np.zeros(shape=10, dtype=int, order='C')
BronchitisSymptoms = np.zeros(shape=11, dtype=int, order='C')
ColdSoreSymptoms = np.zeros(shape=5, dtype=int, order='C')
ConjunctivitisSymptoms = np.zeros(shape=8, dtype=int, order='C')
DiabetesSymptoms = np.zeros(shape=6, dtype=int, order='C')
ChickenpoxSymptoms = np.zeros(shape=8, dtype=int, order='C')
KidneyStonesSymptoms = np.zeros(shape=14, dtype=int, order='C')
MigrainesSymptoms = np.zeros(shape=15, dtype=int, order='C')
PollenAllergySymptoms = np.zeros(shape=10, dtype=int, order='C')
TuberculosisSymptoms = np.zeros(shape=10, dtype=int, order='C')


#-------- Symptom Count and Log ---------#


count = 0
index = 0

for x in Anemia:
    
    if x == 1:
        AnemiaSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#    

count = 0
index = 0

for x in Bronchitis:
    
    if x == 1:
        BronchitisSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in ColdSore:
    
    if x == 1:
        ColdSoreSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in Conjunctivitis:
    
    if x == 1:
        ConjunctivitisSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in Diabetes:
    
    if x == 1:
        DiabetesSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in Chickenpox:
    
    if x == 1:
        ChickenpoxSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in KidneyStones:
    
    if x == 1:
        KidneyStonesSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in Migraines:
    
    if x == 1:
        MigrainesSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in PollenAllergy:
    
    if x == 1:
        PollenAllergySymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  

count = 0
index = 0

for x in Tuberculosis:
    
    if x == 1:
        TuberculosisSymptoms[index] = count
        index = index + 1
        count = count + 1
    elif x == 0:
        count = count+1
#----------------------------------------#  



#--------------------- Anemia Block ------------------------#


NewAnemia = np.copy(Anemia)

stage = 1
count = 0
row = 0

wb = xlsx.Workbook('NewTestWorkbook.xlsx')
ws1 = wb.add_worksheet('Anemia')


while(stage <= AnemiaSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(AnemiaSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    df_matrix = df.as_matrix()
    
    row_old = df_matrix.shape[0]
    
    for x in range(df_matrix.shape[0]):
        ws1.write_row(row, 0, list(df_matrix[x]))
        row += 1
           
    row = row - row_old 
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewAnemia[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#  

        result = algo.Method1(NewAnemia)
        
        ws1.write(row, stage, result)
            
        row += 1

        #---------------------------------#
        
        i = 0
        while(i < stage):
            NewAnemia[NewList[count][i]] = 1
            i = i + 1
        count = count + 1
        
        
    stage = stage + 1
    
wb.close()
print("Anemia done!")   
#----------------------------------------------------------#



'''
#--------------------- Bronchitis Block --------------------------#


NewBronchitis = np.copy(Bronchitis)

stage = 1
count = 0
row = 0


while(stage <= BronchitisSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(BronchitisSymptoms, stage))
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", optimize_write=True, engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Bronchitis', index=False, header=False, startrow=row)
    writer.save()
    
    
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewBronchitis[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewBronchitis)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_csv(writer, sheet_name = 'Bronchitis', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewBronchitis[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------#

print("Bronchitis done!")


#--------------------- Cold Sore Block --------------------------#


NewColdSore = np.copy(ColdSore)

stage = 1
count = 0
row = 0


while(stage <= ColdSoreSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(ColdSoreSymptoms, stage))
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Cold Sore', index=False, header=False, startrow=row)
    writer.save()
    
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewColdSore[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewColdSore)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Cold Sore', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewColdSore[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------#

print("Cold Sore done!")


#--------------------- Conjunctivitis Block --------------------------#


NewConjunctivitis = np.copy(Conjunctivitis)

stage = 1
count = 0
row = 0 

while(stage <= ConjunctivitisSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(ConjunctivitisSymptoms, stage))
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Conjunctivitis', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewConjunctivitis[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewConjunctivitis)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Conjunctivitis', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewConjunctivitis[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------------#

print("Conjunctivitis done!")



#------------------------- Diabetes Block ----------------------------#


NewDiabetes = np.copy(Diabetes)

stage = 1
count = 0
row = 0 

while(stage <= DiabetesSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(DiabetesSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Diabetes', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewDiabetes[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewDiabetes)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Diabetes', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewDiabetes[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------------#


print("Diabetes done!")


#------------------------- Chickenpox Block ----------------------------#


NewChickenpox = np.copy(Chickenpox)

stage = 1
count = 0
row = 0 

while(stage <= ChickenpoxSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(ChickenpoxSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Chickenpox', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewChickenpox[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewChickenpox)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Chickenpox', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewChickenpox[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#----------------------------------------------------------------------#

print("Chickenpox done!")



#------------------------- Kidney Stones Block ----------------------------#


NewKidneyStones = np.copy(KidneyStones)

stage = 1
count = 0
row = 0 

while(stage <= KidneyStonesSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(KidneyStonesSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Kidney Stones', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewKidneyStones[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewKidneyStones)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Kidney Stones', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewKidneyStones[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#------------------------------------------------------------------------#


print("Kidney Stones done!")


#------------------------- Migraines Block ----------------------------#


NewMigraines = np.copy(Migraines)

stage = 1
count = 0
row = 0 

while(stage <= MigrainesSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(MigrainesSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Migraines', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewMigraines[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewMigraines)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Migraines', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewMigraines[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------------#


print("Migraines done!")

#------------------------ Pollen Allergy Block -----------------------#


NewPollenAllergy = np.copy(PollenAllergy)

stage = 1
count = 0
row = 0

while(stage <= PollenAllergySymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(PollenAllergySymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Pollen Allergy', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewPollenAllergy[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewPollenAllergy)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Pollen Allergy', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewPollenAllergy[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------------#

print("Pollen Allergy done!")


#------------------------ Tuberculosis Block -------------------------#


NewTuberculosis = np.copy(Tuberculosis)

stage = 1
count = 0
row = 0 

while(stage <= TuberculosisSymptoms.shape[0]):
    
    count = 0
    
    NewList = list(iter.combinations(TuberculosisSymptoms, stage)) 
    df = pd.DataFrame(NewList)
    book = load_workbook("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx")
    writer = pd.ExcelWriter("/Users/Mason/Documents/Senior-Design-2-Electric-Boogaloo/Test Procedure/Symptom Checker/Test Procedure Data.xlsx", engine="openpyxl")
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name = 'Tuberculosis', index=False, header=False, startrow=row)
    writer.save()
    
    for item in range(len(NewList)):
        i = 0    
        while(i < stage):
            NewTuberculosis[NewList[count][i]] = 0
            i = i + 1
            
        #------ Algorithm Goes Here ------#    
        
        result = algo.Method1(NewTuberculosis)
        ResultsList = []
        ResultsList.append(result)
        df = pd.DataFrame(ResultsList)
        df.to_excel(writer, sheet_name = 'Tuberculosis', index=False, header=False, startrow=row, startcol=stage)
        row += 1
        writer.save()
        
        #---------------------------------#
        
        i = 0    
        while(i < stage):
            NewTuberculosis[NewList[count][i]] = 1
            i = i + 1
        count = count + 1       
        
    stage = stage + 1
    
#---------------------------------------------------------------------#

print("Tuberculosis done!")

'''